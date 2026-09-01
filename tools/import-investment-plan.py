import json
import re
import shutil
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE = Path(r"C:\Users\thalles.silveira\Downloads\Plano de Investimento_10082026.xlsx")
TARGET = ROOT / "data" / "investment-plan-data.js"
PLAN_YEAR = "2026"


def clean(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def number(value):
    if value in (None, ""):
        return ""
    if isinstance(value, (int, float)):
        return value
    text = clean(value).replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return clean(value)


def iso_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = clean(value)
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return None


def region(value):
    text = clean(value)
    mapped = {
        "N": "Norte",
        "NO": "Norte",
        "NE": "Nordeste",
        "CO": "Centro Oeste",
        "SE": "Sudeste",
        "S": "Sul",
    }
    return mapped.get(text.upper(), text)


def split_code_and_name(code_value, name_value):
    raw_name = clean(name_value)
    raw_code = clean(code_value)
    if raw_name:
        match = re.match(r"^(\d{1,8})\s*[.\- ]\s*(.+)$", raw_name)
        if match:
            return match.group(1).zfill(4) if len(match.group(1)) < 4 else match.group(1), match.group(2).strip()
        return raw_code.zfill(4) if raw_code.isdigit() and len(raw_code) < 4 else raw_code, raw_name

    raw_key = clean(code_value)
    key_match = re.match(r"^(\d{1,8})\s*[.\- ]\s*(.+?)(?:\+.+)?$", raw_key)
    if key_match:
        code = key_match.group(1)
        return code.zfill(4) if len(code) < 4 else code, key_match.group(2).strip()
    return raw_code or "0000", raw_key


def main():
    if not SOURCE.exists():
        raise FileNotFoundError(f"Arquivo não encontrado: {SOURCE}")

    wb = load_workbook(SOURCE, data_only=True, read_only=True)
    ws = wb["Plano de Investimento"] if "Plano de Investimento" in wb.sheetnames else wb[wb.sheetnames[0]]

    rows = []
    for index, values in enumerate(ws.iter_rows(min_row=7, max_row=ws.max_row, values_only=True), start=7):
        row_values = list(values)

        def col(number):
            return row_values[number - 1] if number - 1 < len(row_values) else None

        chave_etapa = clean(col(6))
        escopo = clean(col(7))
        if not chave_etapa and not escopo:
            continue
        registro, obra = split_code_and_name(col(5), escopo or chave_etapa)
        if not obra:
            continue

        termino_planejado = iso_date(col(16))
        ano_termino = clean(col(37)) or (termino_planejado or "")[:4]
        rows.append(
            {
                "row": index,
                "registro": registro or "0000",
                "chaveEtapa": chave_etapa or f"Código {registro or '0000'}",
                "obra": obra,
                "tipoUnidade": clean(col(8)),
                "praca": clean(col(9)),
                "uf": clean(col(10)),
                "regiao": region(col(11)),
                "etapaOrdem": number(col(12)),
                "etapa": clean(col(13)),
                "slaDias": number(col(14)),
                "inicioPlanejado": iso_date(col(15)),
                "terminoPlanejado": termino_planejado,
                "terminoReal": iso_date(col(17)),
                "gapPlanejado": number(col(18)),
                "etapaEmAtraso": clean(col(20)),
                "inicioReal": iso_date(col(21)),
                "terminoRealAtual": iso_date(col(22)),
                "terminoFinal": iso_date(col(23)),
                "gapAtual": number(col(24)),
                "indicadorAtraso": clean(col(26)),
                "status": clean(col(27)) or "Planejado",
                "percentualAnterior": clean(col(28)),
                "percentualAtual": clean(col(29)),
                "observacoes": clean(col(30)),
                "dataInclusaoPlano": iso_date(col(31)),
                "dataAjustePlano": iso_date(col(32)),
                "classificacaoObra": clean(col(33)) or "Não informada",
                "tipologiaObra": clean(col(34)) or "Não informada",
                "mesAnoInicioPlanejado": clean(col(35)),
                "mesAnoTerminoPlanejado": clean(col(36)),
                "anoTerminoPlanejado": ano_termino,
                "mesIndicador": clean(col(41)),
                "seAplica": clean(col(42)),
                "totaisAte": clean(col(43)),
                "emAndamento": clean(col(44)),
                "ateODia": clean(col(45)),
                "entregaAntecipada": clean(col(46)),
                "week": clean(col(47)),
            }
        )

    if TARGET.exists():
        backup = TARGET.with_suffix(f".before-plan-20260810-{datetime.now().strftime('%Y%m%d%H%M%S')}.js")
        shutil.copy2(TARGET, backup)

    payload = {
        "source": SOURCE.name,
        "sheet": ws.title,
        "importedAt": datetime.now().isoformat(timespec="seconds"),
        "planYear": PLAN_YEAR,
        "records": rows,
    }
    TARGET.write_text(
        "globalThis.INVESTMENT_PLAN_DATA = "
        + json.dumps(payload, ensure_ascii=False, indent=2)
        + ";\n",
        encoding="utf-8",
    )
    print(json.dumps({"source": SOURCE.name, "sheet": ws.title, "records": len(rows)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
